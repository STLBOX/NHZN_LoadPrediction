from flask import Flask
from openpyxl import load_workbook
from flask_sqlalchemy import SQLAlchemy
import os
import sys
import time
import pandas as pd
from datetime import datetime

WIN = sys.platform.startswith('win')
if WIN:  # 如果是 Windows 系统，使用三个斜线
    prefix = 'sqlite:///'
else:  # 否则使用四个斜线
    prefix = 'sqlite:////'

# 获取当前文件所在的目录
if getattr(sys, 'frozen', False):
    base_dir = os.path.dirname(sys.executable)
elif __file__:
    base_dir = os.path.dirname(__file__)

# 应用
app = Flask(__name__, static_folder=os.path.join(base_dir, 'static'), template_folder=os.path.join(base_dir, 'templates'))
app.config['SQLALCHEMY_DATABASE_URI'] = prefix + os.path.join(base_dir, 'data.db')
db = SQLAlchemy(app)


class User(db.Model):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    user_id = db.Column(db.String(13))
    meter_id = db.Column(db.String(16), primary_key=True)  # 电表的标识
    name = db.Column(db.String(30))
    category = db.Column(db.String(30))


class Usersuqian(db.Model):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    user_id = db.Column(db.String(13))
    meter_id = db.Column(db.String(16), primary_key=True)  # 电表的标识
    name = db.Column(db.String(30))
    category = db.Column(db.String(30))


class Usersuyang(db.Model):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    user_id = db.Column(db.String(13))
    meter_id = db.Column(db.String(16), primary_key=True)  # 电表的标识
    name = db.Column(db.String(30))
    category = db.Column(db.String(30))


class Usersiyang(db.Model):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    user_id = db.Column(db.String(13))
    meter_id = db.Column(db.String(16), primary_key=True)  # 电表的标识
    name = db.Column(db.String(30))
    category = db.Column(db.String(30))


def delete_user_table():
    User.query.delete()  # before 5854
    Usersuqian.query.delete()
    Usersuyang.query.delete()
    Usersiyang.query.delete()
    db.session.commit()



class Power(db.Model):
    '''用户的每天用电功率信息'''
    __table_args__ = (
        db.PrimaryKeyConstraint('meter_id', 'number', 'date'),
    )
    user_id = db.Column(db.String(13))
    meter_id = db.Column(db.String(16), primary_key=True)  # 电表唯一标识
    number = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, primary_key=True)
    power = db.Column(db.Float)


class Powersuqian(db.Model):
    '''用户的每天用电功率信息'''
    __table_args__ = (
        db.PrimaryKeyConstraint('meter_id', 'number', 'date'),
    )
    user_id = db.Column(db.String(13))
    meter_id = db.Column(db.String(16), primary_key=True)  # 电表唯一标识
    number = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, primary_key=True)
    power = db.Column(db.Float)


class Powersuyang(db.Model):
    '''用户的每天用电功率信息'''
    __table_args__ = (
        db.PrimaryKeyConstraint('meter_id', 'number', 'date'),
    )
    user_id = db.Column(db.String(13))
    meter_id = db.Column(db.String(16), primary_key=True)  # 电表唯一标识
    number = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, primary_key=True)
    power = db.Column(db.Float)


class Powersiyang(db.Model):
    '''用户的每天用电功率信息'''
    __table_args__ = (
        db.PrimaryKeyConstraint('meter_id', 'number', 'date'),
    )
    user_id = db.Column(db.String(13))
    meter_id = db.Column(db.String(16), primary_key=True)  # 电表唯一标识
    number = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, primary_key=True)
    power = db.Column(db.Float)


def init_user():
    # data = pd.read_excel("泗洪企业用户.csv")
    # 打开 Excel 文件
    workbook = load_workbook(r'user_table\user_table.xlsx')  # 泗洪县
    # 获取默认的活动工作表
    data = workbook.active
    ids = []
    # 逐行读取数据 客户编号 0	客户名称 1 	行业分类名称 2	电表标识 3
    for row in data.iter_rows(values_only=True):
        if row[3] == '8200000024027938':
            print(row[3])
        meters_id = str(row[3])
        if len(meters_id) < 5:
            continue
        if meters_id in ids:
            continue
        ids.append(meters_id)
        user = User(user_id=str(row[0]), meter_id=meters_id,
                    name=row[1], category=row[2])
        db.session.add(user)
    db.session.commit()
    print(f'初始化完泗洪县的user table, 总记录数据{User.query.count()} ')
    workbook.close()
    # 打开 宿迁Excel 文件
    workbook = load_workbook(r'user_table\user_table_suqian.xlsx')
    # 获取默认的活动工作表
    data = workbook.active
    ids = []
    for row in data.iter_rows(values_only=True):
        meters_id = str(row[3])
        if len(meters_id) < 5:
            continue
        if meters_id in ids:
            continue
        ids.append(meters_id)
        user = Usersuqian(user_id=str(row[0]), meter_id=meters_id,
                    name=row[1], category=row[2])
        db.session.add(user)
    db.session.commit()
    print(f'初始化完宿迁市区的user table, 总记录数据{Usersuqian.query.count()} ')
    workbook.close()
    # 打开 泗阳 Excel 文件
    workbook = load_workbook(r'user_table\user_table_siyang.xlsx')
    # 获取默认的活动工作表
    data = workbook.active
    ids = []
    for row in data.iter_rows(values_only=True):
        meters_id = str(row[3])
        if len(meters_id) < 5:
            continue
        if meters_id in ids:
            continue
        ids.append(meters_id)
        user = Usersiyang(user_id=str(row[0]), meter_id=meters_id,
                    name=row[1], category=row[2])
        db.session.add(user)
    db.session.commit()
    print(f'初始化完泗阳县的user table, 总记录数据{Usersiyang.query.count()} ')
    workbook.close()
    # 打开 沭阳 Excel 文件
    workbook = load_workbook(r'user_table\user_table_suyang.xlsx')
    data = workbook.active
    ids = []
    for row in data.iter_rows(values_only=True):
        meters_id = str(row[3])
        if len(meters_id) < 5:
            continue
        if meters_id in ids:
            continue
        ids.append(meters_id)
        user = Usersuyang(user_id=str(row[0]), meter_id=meters_id,
                    name=row[1], category=row[2])
        db.session.add(user)
    db.session.commit()
    print(f'初始化完沭阳县的user table, 总记录数据{Usersuyang.query.count()} ')
    workbook.close()


if __name__ == "__main__":
    # # 删除用户表 并更新用户表
    # delete_user_table()
    # init_user()
    # time.sleep(5)
    # 打开数据表， 获得其中的所有用户的负荷 数据 并导出
    query_date = datetime(year=2024, month=1, day=18).date()
    meter_power_table = []
    users = [User, Usersuqian, Usersuyang, Usersiyang]
    powers = [Power, Powersuqian, Powersuyang, Powersiyang]

    index = 3
    user = users[index]
    power = powers[index]
    names = ["泗洪县20240118电表功率数据.xlsx", "宿迁市20240118电表功率数据.xlsx",
             "沭阳县20240118电表功率数据.xlsx", "泗阳县20240118电表功率数据.xlsx"]
    for meter_id in user.query.with_entities(user.meter_id).all():
        # 查询 meter_id 数据
        power_data = power.query.\
            with_entities(power.power).\
            filter(power.meter_id == meter_id[0]).\
            filter(power.date == query_date).all()
        if not power_data:
            break
        # 保存在 excel 中 格式 用户名(meter_id) 时刻1(0:00) 时刻2(1:00)...
        res = [meter_id[0]]
        res.extend([data[0] for data in power_data])
        meter_power_table.append(res)
    print(len(meter_power_table))
    # 保存在 excel 中
    columns = ["电表id"] + [str(i) for i in range(0, 24)]
    df = pd.DataFrame(meter_power_table, columns=columns)
    df.to_excel("电表功率数据表格/" + names[index], index=False)