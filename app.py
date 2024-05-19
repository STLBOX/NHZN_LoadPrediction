from mylog import setup_log
import requests
import json
# import pandas as pd
from openpyxl import load_workbook
from flask import Flask, render_template, request, send_file
# from flask import request, redirect, flash, url_for
# from markupsafe import escape
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
# from sqlalchemy.orm import scoped_session, sessionmaker
import os
import sys
import click
from datetime import datetime, timedelta
# from sqlalchemy.exc import IntegrityError
# import copy
from concurrent.futures import ThreadPoolExecutor, as_completed, wait, ALL_COMPLETED
import threading
import time
from urllib.parse import quote
# import logging
# from logging.handlers import TimedRotatingFileHandler
# import re

# 加载全局预测模型
# 获得train_data
from regress_test import *
device = torch.device("cpu")
net = RNNModel(enc_feature_size=8, dec_feature_size=7, hidden_size=32, num_layers=2, dropout=0.2)
net.to(dtype=torch.float32, device=device)
net.load_state_dict(torch.load("model/LSTM.pth"))

# from regress import get_load_data, get_DeepAR
# train_dataset = get_load_data()
# net = get_DeepAR(train_dataset)


logger = setup_log('sx_log.log')

WIN = sys.platform.startswith('win')
if WIN:  # 如果是 Windows 系统，使用三个斜线
    prefix = 'sqlite:///'
else:  # 否则使用四个斜线
    prefix = 'sqlite:////'

# 获取当前文件所在的目录
if getattr(sys, 'frozen', False):
    base_dir = os.path.dirname(sys.executable)
elif __file__:
    base_dir = os.path.dirname(__file__)

# 应用
app = Flask(__name__, static_folder=os.path.join(base_dir, 'static'), template_folder=os.path.join(base_dir, 'templates'))
# 数据库及配置
app.config['SQLALCHEMY_DATABASE_URI'] = prefix + os.path.join(base_dir, 'data.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False  # 关闭对模型修改的监控
app.config['SECRET_KEY'] = 'dev'  # 等同于 app.secret_key = 'dev'
app.config['SQLALCHEMY_POOL_SIZE'] = 20  # 设置连接池大小
os.environ['SQLALCHEMY_SILENCE_UBER_WARNING'] = '1'  # 禁止 warning 消息
app.config.update({
    'SQLALCHEMY_POOL_SIZE': None,
    'SQLALCHEMY_POOL_TIMEOUT': None
})
db = SQLAlchemy(app)


with open('config.txt', 'r', encoding='utf8') as f:
    config_str = f.read()
config_dict = json.loads(config_str)

# 一个json格式记录大屏中的全部数据
DATA = {
    'user_max_power': [],
    'nowpower': 0,
    "power_table_data": {
                        'usernums': "",
                        'pk1max': "",
                        'pk1min': "",
                        'pkmax': "",
                        'pkmin': "",
                        },
    "weather_table_data": {
                        'wtstate': '**',
                        'wtmaxtemp': 0,
                        'wtmintemp': 0,
                        'wtmaxhum': 0,
                        'wtminhum': 0,
                        'wtmaxws': 0,
                        'wtminws': 0
                        },
    'SIMTIME': 0,
    'PAGE': 0,
    'user_top_table': {
        'user_number': ['*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*'],
        'user_id':  ['*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*'],
        'user_name':  ['*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*'],
        'user_cat':  ['*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*', '*'],
        'max_power': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        'max_power_time': ["00:00", "00:00", "00:00", "00:00", "00:00", "00:00", "00:00", "00:00", "00:00", "00:00", "00:00", "00:00", "00:00", "00:00", "00:00", "00:00", "00:00", "00:00", "00:00", "00:00"]},

    "weather_curve": {
        "times": [
            "2023-09-26 00:00:00", "2023-09-26 01:00:00", "2023-09-26 02:00:00", "2023-09-26 03:00:00",
            "2023-09-26 04:00:00", "2023-09-26 05:00:00", "2023-09-26 06:00:00", "2023-09-26 07:00:00",
            "2023-09-26 08:00:00", "2023-09-26 09:00:00", "2023-09-26 10:00:00", "2023-09-26 11:00:00",
            "2023-09-26 12:00:00", "2023-09-26 13:00:00", "2023-09-26 14:00:00", "2023-09-26 15:00:00",
            "2023-09-26 16:00:00", "2023-09-26 17:00:00", "2023-09-26 18:00:00", "2023-09-26 19:00:00",
            "2023-09-26 20:00:00", "2023-09-26 21:00:00", "2023-09-26 22:00:00", "2023-09-26 23:00:00"
        ],
        "timeData": ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12",
                     "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23"],
        "windxData": ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12",
                     "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23"],
        "windsData": ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12",
                     "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23"],
        "tempData": ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12",
                     "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23"],
        "rainData": ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12",
                     "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23"],
        "humData": ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12",
                     "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23"],
    },

    "power_curve": {
        "ptimes": ["2023-09-25 00:00:00", "2023-09-25 01:00:00", "2023-09-25 02:00:00", "2023-09-25 03:00:00",
                    "2023-09-25 04:00:00", "2023-09-25 05:00:00", "2023-09-25 06:00:00", "2023-09-25 07:00:00",
                    "2023-09-25 08:00:00", "2023-09-25 09:00:00", "2023-09-25 10:00:00", "2023-09-25 11:00:00",
                    "2023-09-25 12:00:00", "2023-09-25 13:00:00", "2023-09-25 14:00:00", "2023-09-25 15:00:00",
                    "2023-09-25 16:00:00", "2023-09-25 17:00:00", "2023-09-25 18:00:00", "2023-09-25 19:00:00",
                    "2023-09-25 20:00:00", "2023-09-25 21:00:00", "2023-09-25 22:00:00", "2023-09-25 23:00:00",
                    "2023-09-26 00:00:00", "2023-09-26 01:00:00", "2023-09-26 02:00:00", "2023-09-26 03:00:00",
                    "2023-09-26 04:00:00", "2023-09-26 05:00:00", "2023-09-26 06:00:00", "2023-09-26 07:00:00",
                    "2023-09-26 08:00:00", "2023-09-26 09:00:00", "2023-09-26 10:00:00", "2023-09-26 11:00:00",
                    "2023-09-26 12:00:00", "2023-09-26 13:00:00", "2023-09-26 14:00:00", "2023-09-26 15:00:00",
                    "2023-09-26 16:00:00", "2023-09-26 17:00:00", "2023-09-26 18:00:00", "2023-09-26 19:00:00",
                    "2023-09-26 20:00:00", "2023-09-26 21:00:00", "2023-09-26 22:00:00", "2023-09-26 23:00:00", ],

        "ptimeData": ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13", "14", "15", "16", "17",
                   "18", "19", "20", "21", "22", "23",
                   "0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13", "14", "15", "16", "17",
                   "18", "19", "20", "21", "22", "23"],

        "power": [577, 581, 583, 584, 574, 594,
                625, 671, 720, 801, 851, 873,
                846, 836, 810, 761, 690, 611,
                568, 566, 569, 569, 596, 596,
                577, 581, 583, 584, 574, 594,
                625, 671, 720, 801, 851, 873,
                846, 836, 810, 761, 690, 611,
                568, 566, 569, 569, 596, 596],

        "pred_max": [577, 581, 583, 584, 574, 594,
                   625, 671, 720, 801, 851, 873,
                   846, 836, 810, 761, 690, 611,
                   568, 566, 569, 569, 596, 596,
                   587, 591, 593, 594, 584, 604,
                   635, 681, 730, 821, 871, 903,
                   856, 856, 830, 771, 700, 631,
                   578, 576, 579, 579, 616, 606],

        "pred_min": [577, 581, 583, 584, 574, 594,
                   625, 671, 720, 801, 851, 873,
                   846, 836, 810, 761, 690, 611,
                   568, 566, 569, 569, 596, 596,
                   557, 561, 563, 564, 564, 574,
                   615, 651, 710, 791, 831, 853,
                   836, 816, 800, 751, 670, 601,
                   548, 546, 549, 549, 576, 576],
    }
}

# 初始化的气象数据
weather_data = {
    'temp':[0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
             0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
             0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
             0.00, 0.00, 0.00, 0.00, 0.00, 0.00],
    'rh': [0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
             0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
             0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
             0.00, 0.00, 0.00, 0.00, 0.00, 0.00],
    'wind_dir': ['东南风', '东南风', '东风', '东南风', '东风', '北风',
                 '东南风', '东南风', '东风', '东南风', '东风', '北风',
                 '东南风', '东南风', '东风', '东南风', '东风', '北风',
                 '东南风', '东南风', '东风', '东南风', '东风', '北风'],
    'wind_speed': [0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
             0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
             0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
             0.00, 0.00, 0.00, 0.00, 0.00, 0.00],
    'prec': [0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
             0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
             0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
             0.00, 0.00, 0.00, 0.00, 0.00, 0.00]
}


class User(db.Model):
    '''
    定义一些数据模型将被用于开发负荷预测系统
    关联app：SQlAlchemy   db = SQLAlchemy(app)
    设置： url app.config['SQLALCHEMY_DATABASE_URI']
    创建数据库模型：
    创建数据表和数据库文件
    '''
    '''用户的基本信息'''
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    user_id = db.Column(db.String(13))
    meter_id = db.Column(db.String(16), primary_key=True)  # 电表的标识
    name = db.Column(db.String(30))
    category = db.Column(db.String(30))
    coef = db.Column(db.Float)


class Usersuqian(db.Model):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    user_id = db.Column(db.String(13))
    meter_id = db.Column(db.String(16), primary_key=True)  # 电表的标识
    name = db.Column(db.String(30))
    category = db.Column(db.String(30))
    coef = db.Column(db.Float)

class Usersuyang(db.Model):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    user_id = db.Column(db.String(13))
    meter_id = db.Column(db.String(16), primary_key=True)  # 电表的标识
    name = db.Column(db.String(30))
    category = db.Column(db.String(30))
    coef = db.Column(db.Float)


class Usersiyang(db.Model):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    user_id = db.Column(db.String(13))
    meter_id = db.Column(db.String(16), primary_key=True)  # 电表的标识
    name = db.Column(db.String(30))
    category = db.Column(db.String(30))
    coef = db.Column(db.Float)


class Power(db.Model):
    '''用户的每天用电功率信息'''
    __table_args__ = (
        db.PrimaryKeyConstraint('meter_id', 'number', 'date'),
    )
    user_id = db.Column(db.String(13))
    meter_id = db.Column(db.String(16), primary_key=True)  # 电表唯一标识
    number = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, primary_key=True)
    power = db.Column(db.Float)


class Powersuqian(db.Model):
    '''用户的每天用电功率信息'''
    __table_args__ = (
        db.PrimaryKeyConstraint('meter_id', 'number', 'date'),
    )
    user_id = db.Column(db.String(13))
    meter_id = db.Column(db.String(16), primary_key=True)  # 电表唯一标识
    number = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, primary_key=True)
    power = db.Column(db.Float)


class Powersuyang(db.Model):
    '''用户的每天用电功率信息'''
    __table_args__ = (
        db.PrimaryKeyConstraint('meter_id', 'number', 'date'),
    )
    user_id = db.Column(db.String(13))
    meter_id = db.Column(db.String(16), primary_key=True)  # 电表唯一标识
    number = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, primary_key=True)
    power = db.Column(db.Float)


class Powersiyang(db.Model):
    '''用户的每天用电功率信息'''
    __table_args__ = (
        db.PrimaryKeyConstraint('meter_id', 'number', 'date'),
    )
    user_id = db.Column(db.String(13))
    meter_id = db.Column(db.String(16), primary_key=True)  # 电表唯一标识
    number = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, primary_key=True)
    power = db.Column(db.Float)


class Totalpower(db.Model):
    '''泗洪县所有用户的一天总用电量'''
    __table_args__ = (
        db.PrimaryKeyConstraint('number', 'date'),
    )
    number = db.Column(db.Integer, primary_key=True)  # 时间
    date = db.Column(db.Date, primary_key=True)  # 日期
    total_power = db.Column(db.Float)  # 总功率


class Totalpowersuqian(db.Model):
    '''泗洪县所有用户的一天总用电量'''
    __table_args__ = (
        db.PrimaryKeyConstraint('number', 'date'),
    )
    number = db.Column(db.Integer, primary_key=True)  # 时间
    date = db.Column(db.Date, primary_key=True)  # 日期
    total_power = db.Column(db.Float)  # 总功率


class Totalpowersuyang(db.Model):
    '''泗洪县所有用户的一天总用电量'''
    __table_args__ = (
        db.PrimaryKeyConstraint('number', 'date'),
    )
    number = db.Column(db.Integer, primary_key=True)  # 时间
    date = db.Column(db.Date, primary_key=True)  # 日期
    total_power = db.Column(db.Float)  # 总功率


class Totalpowersiyang(db.Model):
    '''泗洪县所有用户的一天总用电量'''
    __table_args__ = (
        db.PrimaryKeyConstraint('number', 'date'),
    )
    number = db.Column(db.Integer, primary_key=True)  # 时间
    date = db.Column(db.Date, primary_key=True)  # 日期
    total_power = db.Column(db.Float)  # 总功率


class Predictpower(db.Model):
    '''泗洪县所有用户的一天总用电量'''
    __table_args__ = (
        db.PrimaryKeyConstraint('number', 'date'),
    )
    number = db.Column(db.Integer, primary_key=True)  # 时间
    date = db.Column(db.Date, primary_key=True)  # 日期
    predict_power = db.Column(db.Float)  # 预测总功率


class Weather(db.Model):
    '''泗洪地区的气象数据24h'''
    __table_args__ = (
        db.PrimaryKeyConstraint('date', 'hour'),
    )
    date = db.Column(db.Date, primary_key=True)  # 日期
    hour = db.Column(db.Integer, primary_key=True)  # 小时
    temp = db.Column(db.Float)  # 温度
    humidity = db.Column(db.Float)  # 相对湿度
    wind_dir = db.Column(db.String(2))  # 风xiang
    wind_speed = db.Column(db.Float)  # 风速
    prec = db.Column(db.Float)  # 降水量


def format_float_to_str1(float_number):
    return '{:.1f}'.format(float_number)


def format_float_to_str0(float_number):
    return '{:.0f}'.format(float_number)


def format_time_str(start: datetime, length: int):
    res = []
    for i in range(length):
        res.append(start.strftime("%Y-%m-%d %H:%M:%S"))
        start += timedelta(hours=1)
    return res


def get_token():
    # 获取token
    # 解析配置
    url0 = config_dict['url0']
    headers0 = config_dict['headers0']
    data0 = config_dict['data0']
    # post请求，获取token
    request_error = False
    try:
        response = requests.post(url=url0, headers=headers0, data=data0, timeout=5)  # data设置为dict 表单形式传递
        if response.status_code == 200:
            dict_response = json.loads(response.content)  # 解析响应
            logger.info(f"get_token() 获取token成功！ token：{dict_response['access_token']} ")
            response.close()
        else:
            request_error = True
            logger.info(f"get_token() 获取token失败！ 失败原因：响应error_code != 200 ")
    except requests.exceptions.RequestException as e:
        request_error = True
        logger.info(f"获取token失败！ 失败原因：{e} ")
    if request_error:
        dict_response = json.loads('{"access_token": "0ace0809-3e4d-4918-8611-f6581bd16587", '
                                   '"token_type": "bearer", "expires_in": 86399, "scope": "app"}')
        # return -1
    access_token = dict_response['access_token']  # get token
    return access_token


def get_token_weather():
    # 获取token
    # 解析配置
    url0 = config_dict['url0w']
    headers0 = config_dict['headers0w']
    data0 = config_dict['data0w']
    # post请求，获取token
    request_error = False
    try:
        response = requests.post(url=url0, headers=headers0, data=data0, timeout=5)  # data设置为dict 表单形式传递
        if response.status_code == 200:
            dict_response = json.loads(response.content)  # 解析响应
            logger.info(f"get_token_weather() 获取token成功！ token：{dict_response['access_token']} ")
            response.close()
        else:
            request_error = True
            logger.info(f"get_token_weather() 获取token失败！ 失败原因：响应error_code != 200 ")
    except requests.exceptions.RequestException as e:
        request_error = True
        logger.info(f"获取token失败！ 失败原因：{e} ")
    if request_error:
        dict_response = json.loads('{"access_token": "0ace0809-3e4d-4918-8611-f6581bd16587", '
                                   '"token_type": "bearer", "expires_in": 86399, "scope": "app"}')
        # return -1
    access_token = dict_response['access_token']  # get token
    return access_token


def calculate_time(func):
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        run_time = end_time - start_time
        logger.info(f"{func.__name__} 运行时间: {run_time} s")
        return result
    return wrapper


# 用户用电功率数据请求 线程池实现
@calculate_time
def getPowerData_threadingpool(user_num_list, user_list, power_list, totalpower_list):
    # 如果power表K-1时刻的电表数量小于5800*0.9则需要删除重新请求
    date_k1 = datetime.now() - timedelta(days=1)
    max_workers = 2
    executor = ThreadPoolExecutor(max_workers=max_workers)
    file_rto = open(f'user_table/{datetime.now().strftime("%Y%m%d")}_read_time_out.txt', "a")
    file_sr = open(f'user_table/{datetime.now().strftime("%Y%m%d")}_status_error.txt', "a")
    file_nt = open(f'user_table/{datetime.now().strftime("%Y%m%d")}_none_type.txt', "a")
    for user_num, usert, powert, totalpowert in zip(user_num_list, user_list, power_list, totalpower_list):
        if powert.query.filter_by(date=date_k1.date()).count()/24 < user_num*0.7:
            # 删除并获取昨天K-1的功率数据
            powert.query.filter_by(date=date_k1.date()).delete()
            db.session.commit()
            getbefore_num = powert.query.count()
            logger.info(f'getPowerData{powert}_threadingpool()前总记录数据{getbefore_num}')
            if powert.query.filter_by(date=date_k1.date()).first() is None:
                access_token = get_token()  # get token
                # 请求用户的功率数据
                url1 = config_dict['url1']
                headers1 = config_dict['headers1']
                # data1 = config_dict['data1']
                params1 = config_dict['params1']
                headers1['Authorization'] = access_token
                ids = usert.query.with_entities(usert.user_id, usert.meter_id, usert.coef).all()  # 当前的唯一电表ID
                # 循环请求，解析数据并存入数据库，，，可优化为线程池的形式
                # 每个线程分配固定任务
                # 如果请求过程出错，将数据保持为之前数据
                params1['espInformation']['dataDateS'] = date_k1.strftime('%Y-%m-%d')  # "2023-07-21"
                params1['espInformation']['dataDateE'] = date_k1.strftime('%Y-%m-%d')
                def power_task(task_ids):  # task_ids[2] coef
                    for task_id in task_ids:
                        request_error_flag = False
                        params1['espInformation']['meterDevId'] = task_id[1]  # 电表ID
                        json_body = json.dumps(params1)  # 转化为json string type
                        try:
                            # data设置为json格式字符串 以json格式字符串形式传递
                            response = requests.post(url=url1, headers=headers1, data=json_body, timeout=1)
                            if response.status_code == 200:
                                rsp_dict = json.loads(response.content.decode('utf-8'))  # 将响应内容解码为Unicode字符串
                            else:
                                request_error_flag = True
                                logger.info(f"getPowerData() 电表唯一ID:{task_id[1]}，请求失败！ 失败原因：响应 status_code != 200")
                                file_sr.write(task_id[1]+'\n')
                        except requests.exceptions.RequestException as e:
                            request_error_flag = True
                            logger.info(f"getPowerData() 电表唯一ID:{task_id[1]}，请求失败！ 失败原因：{e}")
                            if isinstance(e, requests.exceptions.Timeout):
                                file_rto.write(task_id[1]+'\n')
                        # if request_error_flag:
                        #     logger.info(f"getPowerData() 电表唯一ID:{task_id[1]}，请求失败！ 采用默认值0")
                        #     with open('sim_txt/power.txt', 'r', encoding='utf8') as file:
                        #         rsp_str = file.read()
                        #     rsp_dict = json.loads(rsp_str)
                        # if rsp_dict.get('rtnCode', "0") != '1':
                        # request_error_flag = True
                        # request_error_flag = False
                        date_k2 = (date_k1 - timedelta(days=1)).date()
                        try:
                            for i in range(1, 97, 4):
                                if not request_error_flag:  # 解码保存在数据库中
                                    # 处理特殊情况None
                                    if rsp_dict['rtnCode'] != '1' or len(rsp_dict['dataList'][0]) != 99:
                                        new_power = powert(user_id=task_id[0], meter_id=task_id[1], number=i, date=date_k1, power=0.0)
                                    else:
                                        new_power = powert(user_id=task_id[0], meter_id=task_id[1], number=i, date=date_k1, power=abs(float(rsp_dict['dataList'][0]['v{}'.format(i)]))*task_id[2])
                                else:  # 采用历史K-2数据
                                    # hst_power = powert.query.filter_by(meter_id=task_id[1], date=date_k2, number=i).first()
                                    # new_power = powert(user_id=task_id[0], meter_id=task_id[1], number=i, date=date_k1, power=hst_power.power)
                                    new_power = powert(user_id=task_id[0], meter_id=task_id[1], number=i, date=date_k1, power=0.0)
                                db.session.add(new_power)
                        except Exception as e:
                            logger.info(f"getPowerData_threadingpool() power_task()失败！ 失败原因:{e} ")
                            file_nt.write(task_id[1] + '\n')
                            db.session.rollback()
                        finally:
                            db.session.commit()
                batch_size = int(len(ids)/max_workers)
                get_batch = lambda x: len(ids) if x == (max_workers - 1) else (x+1)*batch_size
                all_tasks = [executor.submit(power_task, ids[i*batch_size:get_batch(i)]) for i in range(0, max_workers)]
                wait(all_tasks, return_when=ALL_COMPLETED)
                add_one_totalpower_recoder(date_k1, powert, totalpowert)
                getafter_num = powert.query.count()
            logger.info(f'getPowerData_threadingpool()后{powert}总记录数据{getafter_num}, 前{getbefore_num}, 电表数量{(getafter_num - getbefore_num)/24}')
    file_nt.close()
    file_sr.close()
    file_rto.close()
    executor.shutdown()

# 用户用电功率数据请求
@calculate_time
def getPowerData():
    '''
    json.loads: str -> dict
    json.dumps: dict -> str
    '''
    date_k1 = datetime.now() - timedelta(days=1)  # 当前的日期
    # 删除K-1时刻的数据
    # Power.query.filter_by(date=date_k1.date()).delete()
    # db.session.commit()

    if Power.query.filter_by(date=date_k1.date()).first() is None:
        access_token = get_token()  # get token
        # 请求用户的功率数据
        url1 = config_dict['url1']
        headers1 = config_dict['headers1']
        # data1 = config_dict['data1']
        params1 = config_dict['params1']
        headers1['Authorization'] = access_token
        ids = User.query.with_entities(User.user_id, User.meter_id).all()  # 当前的唯一电表ID
        # 循环请求，解析数据并存入数据库，，，可优化为线程池的形式
        # 每个线程分配固定任务
        # 如果请求过程出错，将数据保持为之前数据
        # ids = ids[0:15]
        for id in ids:
            request_error_flag = False
            params1['espInformation']['meterDevId'] = id[1]  # 电表ID
            params1['espInformation']['dataDateS'] = date_k1.strftime('%Y-%m-%d')  # "2023-07-21"
            params1['espInformation']['dataDateE'] = date_k1.strftime('%Y-%m-%d')
            json_body = json.dumps(params1)  # 转化为json string type
            try:
                # data设置为json格式字符串 以json格式字符串形式传递
                response = requests.post(url=url1, headers=headers1,  data=json_body, timeout=5)
                if response.status_code == 200:
                    rsp_dict = json.loads(response.content.decode('utf-8'))   # 将响应内容解码为Unicode字符串
                    response.close()
                else:
                    request_error_flag = True
                    logger.info(f"getPowerData() 电表唯一ID:{id[1]}，请求失败！ 失败原因：响应 status_code != 200 ")
            except requests.exceptions.RequestException as e:
                request_error_flag = True
                logger.info(f"getPowerData() 电表唯一ID:{id[1]}，请求失败！ 失败原因：{e} ")

            if request_error_flag:
                with open('sim_txt/power.txt', 'r', encoding='utf8') as file:
                    rsp_str = file.read()
                rsp_dict = json.loads(rsp_str)

            # if rsp_dict.get('rtnCode', "0") != '1':
                # request_error_flag = True
            date_k2 = (date_k1 - timedelta(days=1)).date()
            for i in range(1, 97, 4):
                if not request_error_flag:  # 解码保存在数据库中
                    power = Power(user_id=id[0], meter_id=id[1], number=i, date=date_k1,
                                  power=rsp_dict['dataList'][0]['v{}'.format(i)])
                    db.session.add(power)
                else:  # 采用历史K-2数据
                    hst_power = Power.query.filter_by(meter_id=id[1], date=date_k2, number=i).first()
                    power = Power(user_id=id[0], meter_id=id[1], number=i, date=date_k1, power=hst_power.power)
                    db.session.add(power)
        db.session.commit()


def getSHID(token: str) -> str:
    '''
    Returns: 获得str泗洪县的id
    '''
    # 获得泗洪县的气象站点 id GET
    url2 = config_dict['url2']
    headers2 = config_dict['headers2']
    headers2['Authorization'] = token
    params2 = config_dict['params2']
    request_error = False
    try:
        response = requests.get(url=url2, headers=headers2,  params=params2, timeout=5)
        if response.status_code == 200:
            dict_response = json.loads(response.content)  # 解析响应
            logger.info(f"getSHID()成功！ id为：{dict_response['extra']['id']}")  # 失败后可以采用默认id
            response.close()
        else:
            request_error = True
            logger.info(f"getSHID()失败！ 失败原因：status_code != 200  尝试使用默认id，在sim_txt/shid.txt文件中")  # 失败后可以采用默认id
    except requests.exceptions.RequestException as e:
        request_error = True
        logger.info(f"getSHID()失败！采用默认id! 失败原因:{e} ")  # 失败后可以采用默认id

    if request_error:
        with open('sim_txt/shid.txt', 'r', encoding='utf8') as file:
            rsp_str = file.read()
        dict_response = json.loads(rsp_str)
    shid = dict_response['extra']['id']
    return shid


# 获得天气数据 每天8点更新就行
def getWeatherDataP():
    access_token = get_token_weather()  # get token
    '''
    get 请求的一般做法， 设置requests.get() 的 params参数，字典传递就行， 被附加到URL中
    post 向服务器发送数据， 设置data参数，json传递(json格式传递) 或者 dict(form 表单形式)
    '''
    shid = getSHID(access_token)
    # 逐小时天气预报,当天(0-23， 24h)天气 datetime.now().date()，请求时获得24h  url3
    url3 = config_dict['url3']
    headers3 = config_dict['headers3']
    params3 = config_dict['params3']
    # 修改时间 和 id
    params3['id'] = shid
    headers3['Authorization'] = access_token
    nowtime = datetime.now()
    params3['start_time'] = quote(f'{nowtime.strftime("%Y-%m-%d")} {nowtime.hour+1:02}:00:00')   # K 08:00:00
    params3['end_time'] = quote(f'{(nowtime + timedelta(days=1)).strftime("%Y-%m-%d")} 00:00:00')  # 第二天 00:00:00
    request_error = False
    try:
        response = requests.get(url=url3, headers=headers3,  params=params3, timeout=5)
        if response.status_code == 200:
            logger.info(f"getWeatherDataP() 获取预测值成功！")
            dict_response = json.loads(response.content)  # 解析响应
            response.close()
        else:
            request_error = True
            logger.info(f"getWeatherDataP() 获取预测值失败！ 失败原因：响应error_code != 200 ")
            logger.info(f"response.content: {response.content}")
    except requests.exceptions.RequestException as e:
        # request_error = True
        logger.info(f"first request : getWeatherDataP() 获取预测值失败！ 失败原因：{e} ")
        if isinstance(e, requests.exceptions.Timeout):
            time.sleep(5)
            try:
                response = requests.get(url=url3, headers=headers3, params=params3, timeout=5)
                if response.status_code == 200:
                    logger.info(f"getWeatherDataP() 获取预测值成功！")
                    dict_response = json.loads(response.content)  # 解析响应
                    response.close()
                else:
                    request_error = True
                    logger.info(f"getWeatherDataP() 获取预测值失败！ 失败原因：响应error_code != 200 ")
                    logger.info(f"response.content: {response.content}")
            except requests.exceptions.RequestException as e:
                request_error = True
                logger.info(f"second request : getWeatherDataP() 获取预测值失败！ 失败原因：{e} ")
        else:
            request_error = True
            logger.info(f"first request : getWeatherDataP() 获取预测值失败！ 失败原因：{e} ")

    if request_error:
        with open('sim_txt/prediction.txt', 'r', encoding='utf8') as file:
            rsp_str = file.read()
        dict_response = json.loads(rsp_str)

    # 将预测值保存入数据库
    temp_time = nowtime  # K 07:30:00
    for w in dict_response['extra']:
        # _year = int(w['fb_time'][0:4])
        # _month = int(w['fb_time'][5:7])
        # _day = int(w['fb_time'][8:10])
        # _hour = int(w['fb_time'][11:13])
        # dt = datetime(_year, _month, _day, _hour, 0, 0)
        temp_time = temp_time + timedelta(hours=1)  # K 08:30:00
        # 查询数据库中是否存在指定日期和小时的气象数据,更新预报值 8, 9, 10, ...
        existing_weather = Weather.query.filter(Weather.date == temp_time.date(), Weather.hour == temp_time.hour).first()
        if not request_error:
            if existing_weather is None:
                # 如果不存在，创建新的气象数据
                new_weather = Weather(
                    date=temp_time.date(),
                    hour=temp_time.hour,
                    temp=w['temp'],
                    humidity=w['rh'],
                    wind_dir=w['wind_dir'][:-1],
                    wind_speed=w['wind_speed'],
                    prec=w['prec']
                )
                db.session.add(new_weather)
                db.session.commit()
            else:
                # 如果存在，修改现有的气象数据
                existing_weather.temp = w['temp']
                existing_weather.humidity = w['rh']
                existing_weather.wind_dir = w['wind_dir'][:-1]
                existing_weather.wind_speed = w['wind_speed']
                existing_weather.prec = w['prec']
                db.session.commit()
        else:  # 不存在才替换为历史同期值
            if existing_weather is None:
                hst_weather = Weather.query.filter_by(date=temp_time.date()-timedelta(days=1),
                                                      hour=temp_time.hour).first()
                if hst_weather is None:
                    hst_weather = Weather.query.filter_by(date=temp_time.date() - timedelta(days=2),
                                                          hour=temp_time.hour).first()
                new_weather = Weather(
                    date=temp_time.date(),
                    hour=temp_time.hour,
                    temp=hst_weather.temp,
                    humidity=hst_weather.humidity,
                    wind_dir=hst_weather.wind_dir,
                    wind_speed=hst_weather.wind_speed,
                    prec=hst_weather.prec
                )
                db.session.add(new_weather)
                db.session.commit()


def getWeatherDataN():
    '''
    获得现在数据 不传入时间设置，默认是当前时刻
    Returns:
    '''
    access_token = get_token_weather()  # get token
    shid = getSHID(access_token)
    url5 = config_dict['url5']
    headers5 = config_dict['headers5']
    params5 = config_dict['params5']
    # id
    params5['ids'] = shid
    # token
    headers5['Authorization'] = access_token
    nowtime = datetime.now()
    request_error = False
    try:
        response = requests.get(url=url5, headers=headers5, params=params5, timeout=5)
        if response.status_code == 200:
            dict_response = json.loads(response.content)  # 解析响应
            response.close()
            logger.info(f"getWeatherDataN() 成功！")
        else:
            request_error = True
            logger.info(f"getWeatherDataN() 失败！ 失败原因：响应error_code != 200")

    except requests.exceptions.RequestException as e:
        # request_error = True
        logger.info(f"first request: getWeatherDataN() failed！ the reason：{e}, request again")
        if isinstance(e, requests.exceptions.Timeout):
            time.sleep(5)
            try:
                response = requests.get(url=url5, headers=headers5, params=params5, timeout=5)
                if response.status_code == 200:
                    dict_response = json.loads(response.content)  # 解析响应
                    response.close()
                    logger.info(f"getWeatherDataN() 成功！")
                else:
                    request_error = True
                    logger.info(f"getWeatherDataN() 失败！ 失败原因：响应error_code != 200")
            except requests.exceptions.RequestException as e2:
                request_error = True
                logger.info(f"second request: getWeatherDataN() failed！ the reason：{e2}")
        else:
            request_error = True
            logger.info(f"first request: getWeatherDataN() failed！ the reason：{e}")

    if request_error:
        # 异常处理，采用历史同时刻值
        with open('sim_txt/nowstate.txt', 'r', encoding='utf8') as file:
            rsp_str = file.read()
        dict_response = json.loads(rsp_str)

    # 将当前值保存入数据库  同时将当前的状态态保存到全局的变量
    for w in dict_response['extra']:
        existing_weather = Weather.query.filter(Weather.date == nowtime.date(), Weather.hour == nowtime.hour).first()
        if not request_error:
            if existing_weather is None:
                # 如果不存在，创建新的气象数据
                new_weather = Weather(
                    date=nowtime.date(),
                    hour=nowtime.hour,
                    temp=w['temp'],
                    humidity=w['rh'],
                    wind_dir=w['wind_dir'][:-1],
                    wind_speed=w['wind_speed'],
                    prec=w['prec']
                )
                db.session.add(new_weather)
                db.session.commit()
            else:
                # 如果存在，修改现有的气象数据
                existing_weather.temp = w['temp']
                existing_weather.humidity = w['rh']
                existing_weather.wind_dir = w['wind_dir'][:-1]
                existing_weather.wind_speed = w['wind_speed']
                existing_weather.prec = w['prec']
                db.session.commit()
        else:  #  请求失败，且不存在当前时刻的数据 替换为历史同期值
            if existing_weather is None:
                hst_weather = Weather.query.filter_by(date=nowtime.date() - timedelta(days=1),
                                                      hour=nowtime.hour).first()
                new_weather = Weather(
                    date=nowtime.date(),
                    hour=nowtime.hour,
                    temp=hst_weather.temp,
                    humidity=hst_weather.humidity,
                    wind_dir=hst_weather.wind_dir,
                    wind_speed=hst_weather.wind_speed,
                    prec=hst_weather.prec
                )
                db.session.add(new_weather)
                db.session.commit()

        # 查询数据库
        # w_k = Weather.query.with_entities(Weather.temp, Weather.humidity, Weather.wind_speed, Weather.prec) \
        #     .filter(Weather.date == nowtime.date()) \
        #     .order_by(Weather.hour).all()
        #
        # w_dir = Weather.query.with_entities(Weather.wind_dir) \
        #     .filter(Weather.date == nowtime.date()) \
        #     .order_by(Weather.hour).all()

        # 保存到全局变量  天气的表格 和 天气的曲线
        # 预测气象数据
        # 天气的表格
        # DATA['weather_table_data']['wtstate'] = w['weather_text']
        # DATA['weather_table_data']['wtmaxtemp'] = format_float_to_str1(max([it[0] for it in w_k]))
        # DATA['weather_table_data']['wtmintemp'] = format_float_to_str1(min([it[0] for it in w_k]))
        # DATA['weather_table_data']['wtmaxhum'] = format_float_to_str1(max([it[1] for it in w_k]))
        # DATA['weather_table_data']['wtminhum'] = format_float_to_str1(min([it[1] for it in w_k]))
        # DATA['weather_table_data']['wtmaxws'] = format_float_to_str1(max([it[2] for it in w_k]))  # 最大的风速
        # DATA['weather_table_data']['wtminws'] = format_float_to_str1(min([it[2] for it in w_k]))  # 最小的风速
        #
        # # 天气的曲线 更新
        # DATA["weather_curve"]["windxData"] = [it[0] for it in w_dir]
        # DATA["weather_curve"]["tempData"] = [format_float_to_str1(it[0]) for it in w_k]
        # DATA["weather_curve"]["humData"] = [format_float_to_str1(it[1]) for it in w_k]
        # DATA["weather_curve"]["windsData"] = [format_float_to_str1(it[2]) for it in w_k]
        # DATA["weather_curve"]["rainData"] = [format_float_to_str1(it[3]) for it in w_k]


def init_db():
    db.drop_all()
    db.create_all()
    db.session.execute('VACUUM;')


def init_user():
    # data = pd.read_excel("泗洪企业用户.csv")
    # 打开 Excel 文件
    workbook = load_workbook(r'user_table\user_table.xlsx')  # 泗洪县
    # 获取默认的活动工作表
    data = workbook.active
    ids = []
    # 逐行读取数据 客户编号 0	客户名称 1 	行业分类名称 2	电表标识 3   系数 4
    for row in data.iter_rows(values_only=True):
        meters_id = str(row[3])
        if len(meters_id) < 5:
            continue
        if meters_id in ids:
            continue
        ids.append(meters_id)
        user = User(user_id=str(row[0]), meter_id=meters_id,
                    name=row[1], category=row[2], coef=row[4])
        db.session.add(user)
    db.session.commit()
    logger.info(f'初始化完泗洪县的user table, 总记录数据{User.query.count()} ')
    workbook.close()
    # 打开 宿迁Excel 文件
    workbook = load_workbook(r'user_table\user_table_suqian.xlsx')
    # 获取默认的活动工作表
    data = workbook.active
    ids = []
    for row in data.iter_rows(values_only=True):
        meters_id = str(row[3])
        if len(meters_id) < 5:
            continue
        if meters_id in ids:
            continue
        ids.append(meters_id)
        user = Usersuqian(user_id=str(row[0]), meter_id=meters_id,
                    name=row[1], category=row[2], coef=row[4])
        db.session.add(user)
    db.session.commit()
    logger.info(f'初始化完宿迁市区的user table, 总记录数据{Usersuqian.query.count()} ')
    workbook.close()
    # 打开 泗阳 Excel 文件
    workbook = load_workbook(r'user_table\user_table_siyang.xlsx')
    # 获取默认的活动工作表
    data = workbook.active
    ids = []
    for row in data.iter_rows(values_only=True):
        meters_id = str(row[3])
        if len(meters_id) < 5:
            continue
        if meters_id in ids:
            continue
        ids.append(meters_id)
        user = Usersiyang(user_id=str(row[0]), meter_id=meters_id,
                    name=row[1], category=row[2], coef=row[4])
        db.session.add(user)
    db.session.commit()
    logger.info(f'初始化完泗阳县的user table, 总记录数据{Usersiyang.query.count()} ')
    workbook.close()
    # 打开 沭阳 Excel 文件
    workbook = load_workbook(r'user_table\user_table_suyang.xlsx')
    data = workbook.active
    ids = []
    for row in data.iter_rows(values_only=True):
        meters_id = str(row[3])
        if len(meters_id) < 5:
            continue
        if meters_id in ids:
            continue
        ids.append(meters_id)
        user = Usersuyang(user_id=str(row[0]), meter_id=meters_id,
                          name=row[1], category=row[2], coef=row[4])
        db.session.add(user)
    db.session.commit()
    logger.info(f'初始化完沭阳县的user table, 总记录数据{Usersuyang.query.count()} ')
    workbook.close()


# 初始化数据库连接池
def init_pool():
    conn = db.engine.connect()
    conn.close()


def init_one_user_power(user_ids, power):
    # 初始化的电力数据 数据库仅保存最近3天的数据，在这里初始化，全部相同  MW
    nums = [1, 5, 9, 13, 17, 21, 25, 29, 33, 37, 41, 45, 49, 53, 57, 61, 65, 69, 73, 77, 81, 85, 89, 93]
    power_data = [0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
                  0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
                  0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
                  0.00, 0.00, 0.00, 0.00, 0.00, 0.00]
    # 使用 Session 来创建会话对象
    # session = db.session
    # with db_write_lock:
    date = datetime.now() - timedelta(days=2)
    for user_id in user_ids:
        try:
            for i in range(len(nums)):
                new_power = power(user_id=user_id[1], meter_id=user_id[0], number=nums[i], date=date.date(), power=power_data[i]*user_id[2])
                db.session.add(new_power)
        except Exception as e:
             db.session.rollback()
             logger.info(f'init_one_user_power()错误， rollback()数据库 {e}')
        finally:
            db.session.commit()
            # session.close()
        # session.add(power)
        # 最后记得关闭会话
        # session.close()
        # try:
        #
        # except Exception as e:
        #     # 加入数据库commit提交失败，必须回滚！！！
        #     db.session.rollback()
        #     print(f"{e}")


def add_totalpower_recoder(date: datetime):
    power_list = [Power, Powersuqian, Powersuyang, Powersiyang]
    total_power_list = [Totalpower, Totalpowersuqian, Totalpowersuyang, Totalpowersiyang]
    for power, totalpower in zip(power_list, total_power_list):
        # 查询一天逐时功率的总和  24  24*1
        p = db.session.query(func.sum(power.power)) \
              .where(power.date == date.date()) \
              .group_by(power.number) \
              .order_by(power.number).all()
        for i in range(24):
            new_totalpower = totalpower(
                number=i,
                date=date.date(),
                total_power=p[i][0],
            )
            db.session.add(new_totalpower)
        db.session.commit()


def add_one_totalpower_recoder(date: datetime, power, totalpower):
        # 查询一天逐时功率的总和  24  24*1  单位转化为MW
        p = db.session.query(func.sum(power.power)) \
              .where(power.date == date.date()) \
              .group_by(power.number) \
              .order_by(power.number).all()
        for i in range(24):
            new_totalpower = totalpower(
                number=i,
                date=date.date(),
                total_power=p[i][0]/1000
            )
            db.session.add(new_totalpower)
        db.session.commit()


def save_predictpower(data: list, date: datetime):
    for i in range(24):
        new_predictpower = Predictpower(
            number=i,
            date=date.date(),
            predict_power=data[i],
        )
        db.session.add(new_predictpower)
    db.session.commit()


# @calculate_time
def init_power(user_num_list, user_list, power_list):
    # Power.query.delete()
    # db.session.commit()
    date = datetime.now() - timedelta(days=2)
    # 判断数据是否存在，存在就跳过初始化
    max_workers = 2
    executor = ThreadPoolExecutor(max_workers=max_workers)
    for user_num, user, power in zip(user_num_list, user_list, power_list):
        if power.query.filter_by(date=(datetime.now() - timedelta(days=2)).date()).count() / 24 < user_num * 0.5:
            ids = user.query.with_entities(user.meter_id, user.user_id, user.coef).all()  # 当前的唯一电表ID  用户ID 系数
            logger.info(f'the number of total mentors are {len(ids)}')
            if power.query.filter_by(date=date.date()).first() is None:
                # db.session.close()  # 关闭当前会话
                # init_one_user_power(ids[0])
                batch_size = int(len(ids) / max_workers)
                get_batch = lambda x: len(ids) if x == (max_workers - 1) else (x + 1) * batch_size
                all_tasks = [executor.submit(init_one_user_power, ids[i * batch_size:get_batch(i)], power) for i in range(0, max_workers)]
                wait(all_tasks, return_when=ALL_COMPLETED)
                # db.session.commit()
        logger.info(f'初始化完成K-2时的power and totalpower table, 总记录数据{power.query.count()}')
    executor.shutdown()


def init_predict_power():
    # k-2时刻初始化
    # datek2 = datetime.now() - timedelta(days=2)
    # if Predictpower.query.filter_by(date=datek2.date()).first() is None:
    #     # 查询数据并赋值
    #     powerk2 = Totalpower.query.with_entities(Totalpower.total_power) \
    #         .filter(Totalpower.date == datek2.date()) \
    #         .order_by(Totalpower.number).all()
    #
    #     save_predictpower(data=[p[0] for p in powerk2], date=datek2)
    # k-1时刻初始化
    datek1 = datetime.now() - timedelta(days=1)
    if Predictpower.query.filter_by(date=datek1.date()).first() is None:
        powerk1 = Totalpower.query.with_entities(Totalpower.total_power) \
            .filter(Totalpower.date == datek1.date()) \
            .order_by(Totalpower.number).all()
        save_predictpower(data=[p[0] for p in powerk1], date=datek1)


def init_weather_partial(
    url: str,
    headers: dict,
    params: dict,
    start_time: datetime,
    end_time: datetime,
    init_nowday: bool = False):
    '''
    初始化指定时间范围内的历史数据库
    Args:
        start_time:
        end_time:

    Returns: None
    '''
    params['start_time'] = quote(f'{start_time.strftime("%Y-%m-%d")} {start_time.hour:02}:{start_time.minute:02}:00')
    params['end_time'] = quote(f'{end_time.strftime("%Y-%m-%d")} {end_time.hour:02}:{end_time.minute:02}:00')
    request_flag = False
    try:
        response = requests.get(url=url, headers=headers,  params=params, timeout=5)
        if response.status_code == 200:
            dict_response = json.loads(response.content)  # 解析响应
            response.close()
            logger.info(f'init_weather_partial() 成功')
        else:
            request_flag = True
            logger.info(f'init_weather_partial() 失败， 失败原因: status_code != 200')
    except requests.exceptions.RequestException as e:
        request_flag = True
        logger.info(f'init_weather_partial() 失败， 失败原因{e}')
    # 请求出错的话
    if request_flag:
        logger.info(f'request_flag:{request_flag}, 采用默认值from sim_txt/hst.txt')
        with open('sim_txt/hst.txt', 'r', encoding='utf8') as file:
            rsp_str = file.read()
        dict_response = json.loads(rsp_str)

    # 解析数据，保存结果，失败的话选择默认值
    # 间隔 12 [0 ,12 ,24 ... ]
    kernel_data = dict_response['extra'][params['id']]
    index = 0
    len_data = int(len(kernel_data)/12) + 1  # 24
    logger.info(f'init_weather_partial() 时获取的数据的数量为{len_data}, request_flag,{request_flag}')
    # 检查数据是够一致，条目数量为24
    # assert(len_data == 24), "数据不一致，返回数据不满足或者超过24h"
    if len_data != 24:
        request_flag = True  # 数据不一致也采用默认的数据 全部为0
    while index < len_data:
        temp_time = start_time + timedelta(hours=index)  # 第index个小时
        # 检查数据是否存在
        exist = Weather.query.filter(Weather.date == temp_time.date(), Weather.hour == temp_time.hour).first()  # 不存在才初始
        if exist is None:
            if not request_flag:  # 请求成功
                temp_data = kernel_data[index*12]
                new_weather = Weather(
                    date=temp_time.date(),
                    hour=temp_time.hour,
                    temp=temp_data['temp'],
                    humidity=temp_data['rh'],
                    wind_dir=temp_data['wind_dir'][:-1],
                    wind_speed=temp_data['wind_speed'],
                    prec=temp_data['prec'],
                )
                db.session.add(new_weather)
                if init_nowday:
                    now_weather = Weather(
                        date=(temp_time.date() + timedelta(days=1)),
                        hour=temp_time.hour,
                        temp=temp_data['temp'],
                        humidity=temp_data['rh'],
                        wind_dir=temp_data['wind_dir'][:-1],
                        wind_speed=temp_data['wind_speed'],
                        prec=temp_data['prec'],
                    )
                    db.session.add(now_weather)
            else:  # 请求失败
                new_weather = Weather(
                    date=temp_time.date(),
                    hour=temp_time.hour,
                    temp=weather_data['temp'][index],
                    humidity=weather_data['rh'][index],
                    wind_dir=weather_data['wind_dir'][index][:-1],
                    wind_speed=weather_data['wind_speed'][index],
                    prec=weather_data['prec'][index],
                )
                db.session.add(new_weather)
                if init_nowday:
                    now_weather = Weather(
                        date=(temp_time.date()+timedelta(days=1)),
                        hour=temp_time.hour,
                        temp=weather_data['temp'][index],
                        humidity=weather_data['rh'][index],
                        wind_dir=weather_data['wind_dir'][index][:-1],
                        wind_speed=weather_data['wind_speed'][index],
                        prec=weather_data['prec'][index],
                    )
                    db.session.add(now_weather)
        index += 1
    db.session.commit()


def init_weather():
    # 初始化气象数据
    # 尝试request历史数据初始化，如果失败，则采用默认值，冷启动的方式运行程序
    # 实际运行时，每间隔一个h请求当前实时数据，并保存
    access_token = get_token_weather()
    shid = getSHID(access_token)

    url4 = config_dict['url4']
    headers4 = config_dict['headers4']
    headers4['Authorization'] = access_token
    params4 = config_dict['params4']
    # 修改时间 和 id
    params4['id'] = shid

    nowtime = datetime.now() - timedelta(days=1)
    start_time = datetime(nowtime.year, nowtime.month, nowtime.day, 0, 0, 0)
    end_time = datetime(nowtime.year, nowtime.month, nowtime.day, 23, 30, 0)
    init_weather_partial(url4, headers4, params4, start_time, end_time, True)
    # 前2天的数据  K-2
    # nowtime = datetime.now() - timedelta(days=2)
    # start_time = datetime(nowtime.year, nowtime.month, nowtime.day, 0, 0, 0)
    # end_time = datetime(nowtime.year, nowtime.month, nowtime.day, 23, 30, 0)
    # init_weather_partial(url4, headers4, params4, start_time, end_time)
    logger.info(f'气象数据 K K-1 初始化完成 ')


# 仅保留3个月内的数据，之前的全部删除
def delete_hst_data():
    logger.info(f"delete_hst_data()删除前Power总记录数量{Power.query.count()}")
    three_days_ago = datetime.now() - timedelta(days=4)  # 三天前
    Power.query.filter(Power.date < three_days_ago.date()).delete()
    Powersuqian.query.filter(Powersuqian.date < three_days_ago.date()).delete()
    Powersuyang.query.filter(Powersuyang.date < three_days_ago.date()).delete()
    Powersiyang.query.filter(Powersiyang.date < three_days_ago.date()).delete()

    three_years_ago = datetime(datetime.now().year - 3, 1, 1, 0, 0, 0)  # 气象数据3年前
    Weather.query.filter(Weather.date < three_years_ago.date()).delete()
    Totalpower.query.filter(Totalpower.date < three_years_ago.date()).delete()
    Totalpowersuqian.query.filter(Totalpowersuqian.date < three_years_ago.date()).delete()
    Totalpowersuyang.query.filter(Totalpowersuyang.date < three_years_ago.date()).delete()
    Totalpowersiyang.query.filter(Totalpowersiyang.date < three_years_ago.date()).delete()
    db.session.commit()
    logger.info(f"delete_hst_data()删除后Power总记录数量{Power.query.count()}")
    db.session.execute('VACUUM;')


def delete_user_table():
    logger.info(f"重置user_table")
    User.query.delete()
    Usersuqian.query.delete()
    Usersuyang.query.delete()
    Usersiyang.query.delete()
    db.session.commit()
    # db.session.execute('VACUUM;')


def predict_power():
    nowtime = datetime.now()  # - timedelta(days=4)
    # search data keep in the global obj
    # 准备数据 by 查询数据库
    # 历史气象数据(k - 1) 24*4, 预测气象数据(k) 24*4, 历史总功率数据(k - 1)  24*1
    w_k1 = Weather.query.with_entities(Weather.date, Weather.hour, Weather.temp, Weather.humidity, Weather.wind_speed, Weather.prec) \
        .filter(Weather.date == (nowtime - timedelta(days=1)).date())\
        .order_by(Weather.hour).all()
    # 预测气象数据
    w_k = Weather.query.with_entities(Weather.date, Weather.hour, Weather.temp, Weather.humidity, Weather.wind_speed, Weather.prec) \
        .filter(Weather.date == nowtime.date())\
        .order_by(Weather.hour).all()
    w_dir = Weather.query.with_entities(Weather.wind_dir) \
        .filter(Weather.date == nowtime.date())\
        .order_by(Weather.hour).all()
    # 每个用户每天的24时刻的瞬时功率
    user_sum_power = Power.query.with_entities(Power.user_id, Power.number, func.sum(Power.power).label('sum_power'))\
        .filter(Power.date == (nowtime - timedelta(days=1)).date()) \
        .group_by(Power.user_id, Power.number).subquery()
    # 对用户的瞬时功率排序,找出每个用户的最大瞬时功率  5367  注意是用户 同一用户 可能存在多个 电表
    user_max_power = db.session.query(user_sum_power.c.user_id, user_sum_power.c.number, func.max(user_sum_power.c.sum_power).label('max_power')) \
        .group_by(user_sum_power.c.user_id) \
        .order_by(func.max(user_sum_power.c.sum_power).desc()).all()
    DATA['user_max_power'] = user_max_power[0:30]  # 用户功率保持为KW单位
    # 查询一天逐时功率的总和  24  24*1
    # p_k1 = db.session.query(func.sum(user_sum_power.c.sum_power).label('total_power')) \
    #     .group_by(user_sum_power.c.number) \
    #     .order_by(user_sum_power.c.number).all()
    p_k1 = Totalpower.query.with_entities(Totalpower.total_power)\
                     .filter(Totalpower.date == (nowtime - timedelta(days=1)).date())\
                     .order_by(Totalpower.number).all()
    # 获得预测数据
    yesterday = nowtime - timedelta(days=1)
    start_str = datetime(yesterday.year, yesterday.month, yesterday.day, 0, 0, 0).strftime("%Y-%m-%d %H:%M:%S")
    end_str = datetime(nowtime.year, nowtime.month, nowtime.day, 23, 0, 0).strftime("%Y-%m-%d %H:%M:%S")
    # 构造数据
    # power_sub = [[itemp[0], itemw[0].month, itemw[0].day, itemw[0].weekday(), itemw[1],
    #               itemw[2], itemw[3], itemw[4]] for itemp, itemw in zip(power_day, weather_day)]
    # p1, p50, p99 = get_next_day_power(w_k1, p_k1, w_k, net, device)  # return is list
    p1, p50, p99 = get_next_day_power_k1(w_k1, p_k1, w_k)
    # p1, p50, p99 = net.get_next_day_power(w_k1, p_k1, w_k, train_dataset, start_str, end_str)
    # 保存预测结果到database
    # 检查是否存在预测结果
    if Predictpower.query.filter_by(date=nowtime.date()).first() is None:
        save_predictpower(data=p50, date=nowtime)

    # 保持数据到全局对象  1*48   气象数据曲线： 也可以更新
    DATA["power_curve"]["ptimes"] = format_time_str(datetime(yesterday.year, yesterday.month, yesterday.day, 0, 0, 0), 48)
    DATA["power_curve"]["power"] = [format_float_to_str1(it[0]) for it in p_k1] + [format_float_to_str1(it) for it in p50]
    DATA["power_curve"]["pred_max"] = [format_float_to_str1(it[0]) for it in p_k1] + [format_float_to_str1(it) for it in p99]
    DATA["power_curve"]["pred_min"] = [format_float_to_str1(it[0]) for it in p_k1] + [format_float_to_str1(it) for it in p1]

    # 保持数据到全局对象  1*24   气象数据曲线： 也可以更新
    DATA["weather_curve"]["times"] = format_time_str(datetime(nowtime.year, nowtime.month, nowtime.day, 0, 0, 0), 24)
    # 风向
    DATA["weather_curve"]["windxData"] = [it[0] for it in w_dir]
    DATA["weather_curve"]["tempData"] = [format_float_to_str1(it[2]) for it in w_k]
    DATA["weather_curve"]["humData"] = [format_float_to_str1(it[3]) for it in w_k]
    DATA["weather_curve"]["windsData"] = [format_float_to_str1(it[4]) for it in w_k]
    DATA["weather_curve"]["rainData"] = [format_float_to_str1(it[5]) for it in w_k]

    DATA['power_table_data']['usernums'] = format_float_to_str0(len(user_max_power))
    DATA['power_table_data']['pk1max'] = format_float_to_str1(max(p_k1)[0])
    DATA['power_table_data']['pk1min'] = format_float_to_str1(min(p_k1)[0])
    DATA['power_table_data']['pkmax'] = format_float_to_str1(max(p50))
    DATA['power_table_data']['pkmin'] = format_float_to_str1(min(p50))
    _hour = datetime.now().hour
    DATA['nowpower'] = DATA["power_curve"]["power"][24 + _hour]

    DATA['weather_table_data']['wtstate'] = '阴'
    DATA['weather_table_data']['wtmaxtemp'] = format_float_to_str1(max([it[2] for it in w_k]))
    DATA['weather_table_data']['wtmintemp'] = format_float_to_str1(min([it[2] for it in w_k]))
    DATA['weather_table_data']['wtmaxhum'] = format_float_to_str1(max([it[3] for it in w_k]))
    DATA['weather_table_data']['wtminhum'] = format_float_to_str1(min([it[3] for it in w_k]))
    DATA['weather_table_data']['wtmaxws'] = format_float_to_str1(max([it[4] for it in w_k]))  # 最大的风速
    DATA['weather_table_data']['wtminws'] = format_float_to_str1(min([it[4] for it in w_k]))  # 最小的风速


# 1 hour
@app.route('/update_nav', methods=['GET'])
def update_nav():
    '''
    导航： 时间  nowtime 和预测实时 nowpower 功率显示分辨率(1h)
    Returns:
    '''
    _hour = datetime.now().hour
    # _hour = datetime.now().second % 24
    data_nav = {"nowpower": DATA["power_curve"]["power"][24 + _hour]}  # 0-23
    return data_nav


# 1 day
@app.route('/update_power_table', methods=['GET'])
def update_power_table():
    return DATA['power_table_data']


# 1 day
@app.route('/update_power_curve', methods=['GET'])
def update_power_curve():
    return DATA['power_curve']
    # DATAc = copy.deepcopy(DATA['power_curve'])
    # DATAc["power"][DATA['SIMTIME']] = 100000
    # DATAc["pred_max"][DATA['SIMTIME']] = 100000
    # DATAc["pred_min"][DATA['SIMTIME']] = 100000
    # return DATAc


# 1 hour
@app.route('/update_weather_table', methods=['GET'])
def update_weather_table():
    return DATA['weather_table_data']


def back_update_weather_table():
    getWeatherDataN()


# 1 hour
@app.route('/update_weather_curve', methods=['GET'])
def update_weather_curve():
    hour_now = datetime.now().hour
    # hour_now = DATA['SIMTIME']
    start = min(hour_now, 12)
    DATAc = {
        "times": DATA['weather_curve']['times'][start:start + 12],
        "timeData": DATA['weather_curve']['timeData'][start:start + 12],
        "windxData": DATA['weather_curve']['windxData'][start:start + 12],
        "windsData": DATA['weather_curve']['windsData'][start:start + 12],
        "tempData": DATA['weather_curve']['tempData'][start:start + 12],
        "rainData": DATA['weather_curve']['rainData'][start:start + 12],
        "humData": DATA['weather_curve']['humData'][start:start + 12],
    }
    # DATA['SIMTIME'] = (DATA['SIMTIME'] + 1) % 24  # update SIM_TIME
    return DATAc


def format_time(hour, minute):
    formatted_hour = str(hour).zfill(2)  # 将小时转换为两位数字符串
    formatted_minute = str(minute).zfill(2)  # 将分钟转换为两位数字符串
    return "{}:{}".format(formatted_hour, formatted_minute)


def remove_prefix(input_string):
    prefixes = ["泗洪", "泗洪县"]
    for prefix in prefixes:
        if input_string.startswith(prefix):
            return input_string[len(prefix):]
    return input_string


@app.route('/update_user_top_table', methods=['GET'])
def update_user_top_table():
    # 传给前端用电量前20个用户的信息，自己滚动
    user_max_power = DATA['user_max_power']
    # 找出每个用户的信息
    # start = DATA['PAGE']*5
    # end = (DATA['PAGE'] + 1) * 5
    for row in range(20):
        DATA['user_top_table']['user_number'][row] = row + 1
        DATA['user_top_table']['user_id'][row] = user_max_power[row][0]
        # 查询用户的名字
        user_info = User.query.with_entities(User.name, User.category).filter(User.user_id == user_max_power[row][0]).first()
        DATA['user_top_table']['user_name'][row] = remove_prefix(user_info[0])
        DATA['user_top_table']['user_cat'][row] = user_info[1]
        DATA['user_top_table']['max_power'][row] = format_float_to_str1(user_max_power[row][2])
        DATA['user_top_table']['max_power_time'][row] = format_time(hour=int((user_max_power[row][1]-1)/4), minute=0)
    # DATA['PAGE'] = (DATA['PAGE']+1) % 4
    return DATA['user_top_table']


# 执行定时更新，请求由后台驱动  6点开始更新
# 1.6:05获取前一天的所有用户功率信息，以及预测气象数据  back_update_power_table()
# 2.每个小时的10分，获取当前气象数据  back_update_weather_table()
def update_all():
    while True:
        now = datetime.now()
        if now.minute == 5 or now.minute == 4 or now.minute == 3:
            # print("back_update_weather_table() 成功")
            back_update_weather_table()
            logger.info("back_update_weather_table() 成功")
        if (now.hour == 6) and (now.minute == 10 or now.minute == 11 or now.minute == 12):   # now.hour == 6 &
            # print("back_update_power_table() 成功")
            # back_update_power_table()
            logger.info("back_update_power_table() ")
            delete_hst_data()
        time.sleep(60*3)


def update_power_thread(user_num_list, user_list, power_list, totalpower_list):
    while True:
        now = datetime.now()
        if (now.hour == 5) and (now.minute == 1 or now.minute == 2 or now.minute == 3):
            getPowerData_threadingpool(user_num_list, user_list, power_list, totalpower_list)
            logger.info("getPowerData_threadingpool() 结束")
            delete_hst_data()
            logger.info("delete_hst_data() 结束")
        time.sleep(60*3)


def update_weather_thread():
    while True:
        now = datetime.now()
        if now.minute == 5 or now.minute == 4 or now.minute == 3:  # 每一个小时更新一次
            # print("back_update_weather_table() 成功")
            getWeatherDataN()
            logger.info("getWeatherDataN() 结束")
        if (now.hour == 8) and (now.minute == 10 or now.minute == 11 or now.minute == 12):  # now.hour == 6 &
            getWeatherDataP()   # [8, 9, 10, 11 ... 23] 气象预测数据
            predict_power()
            logger.info("getWeatherDataP() 和 predict_power() 结束")
        time.sleep(60 * 3)


@app.cli.command()  # 注册为命令，可以传入 name 参数来自定义命令
@click.option('--drop', is_flag=True, help='Create after drop.')  # 设置选项
def initdb(drop=False):
    """Initialize the database."""
    if drop:  # 判断是否输入了选项
        db.drop_all()
        click.echo('delete database.')  # 输出提示信息
    db.create_all()
    click.echo('Initialized database.')  # 输出提示信息


# 设置静态资源的缓存头部
# @app.route('/static/<path:filename>')
# def static_file(filename):
#     response = send_file(filename, cache_timeout=86400)  # 设置缓存时间，单位是秒
#     return response


@app.route('/')
def index():
    return render_template('main.html')


def get_time_str(item, hour):
    return datetime(item.year, item.month, item.day, hour, 0, 0).strftime('%Y-%m-%d %H:%M')


def query_st_ed(st, ed):
    # 查询日期的最小值和最大值
    datetime_min = db.session.query(func.min(Totalpower.date)).scalar()
    datetime_max = db.session.query(func.max(Totalpower.date)).scalar()
    if ed < st:
        ed, st = datetime_max, datetime_max
        return query_st_ed(st, ed)
    if st < datetime_min or ed > datetime_max:
        if st < datetime_min:
            st = datetime_min
        if ed > datetime_max:
            ed = datetime_max
        return query_st_ed(st, ed)
    # 查询st ed之间的数据,排序，并构造[]list返回
    result_sh = Totalpower.query.with_entities(Totalpower.date, Totalpower.number, Totalpower.total_power)\
                .filter(Totalpower.date >= st, Totalpower.date <= ed)\
                .order_by(Totalpower.date, Totalpower.number)
    if result_sh.first() is None:
        ed, st = (datetime.now() - timedelta(days=1)).date(), (datetime.now() - timedelta(days=1)).date()
        return query_st_ed(st, ed)
    # 查询结果共用一个x
    curve_time = []
    curve_sh = []
    for item in result_sh:
        curve_time.append(datetime(item[0].year, item[0].month, item[0].day, item[1], 0, 0).strftime('%Y-%m-%d %H:%M:%S'))
        curve_sh.append(round(item[2], 1))

    result_suqian = Totalpowersuqian.query.with_entities(Totalpowersuqian.total_power)\
                    .filter(Totalpowersuqian.date >= st, Totalpowersuqian.date <= ed)\
                    .order_by(Totalpowersuqian.date, Totalpowersuqian.number)
    curve_sq = [round(item[0], 1) for item in result_suqian]
    result_suyang = Totalpowersuyang.query.with_entities(Totalpowersuyang.total_power)\
                    .filter(Totalpowersuyang.date >= st, Totalpowersuyang.date <= ed)\
                    .order_by(Totalpowersuyang.date, Totalpowersuyang.number)
    curve_suyang = [round(item[0], 1) for item in result_suyang]
    result_siyang = Totalpowersiyang.query.with_entities(Totalpowersiyang.total_power)\
                    .filter(Totalpowersiyang.date >= st, Totalpowersiyang.date <= ed)\
                    .order_by(Totalpowersiyang.date, Totalpowersiyang.number)
    curve_siyang = [round(item[0], 1) for item in result_siyang]
    # 预测的power
    pred_sh = Predictpower.query.with_entities(Predictpower.predict_power)\
                    .filter(Predictpower.date >= st, Predictpower.date <= ed)\
                    .order_by(Predictpower.date, Predictpower.number)
    curve_pred_sh = [round(item[0], 1) for item in pred_sh]
    # 重新构造pie_data,数据为最新的一天的对应时刻的数据

    pie_data = [curve_sh[0], curve_sq[0], curve_suyang[0], curve_siyang[0]]
    # curve_time[0] = '2023-11-24 01:00:00'
    # 查询区间内的最高功率取第一个各值 最低功率及对应的天气

    max_info = []
    for power, curve_data in zip([Totalpower, Totalpowersuqian, Totalpowersuyang, Totalpowersiyang],
                                 [curve_sh, curve_sq, curve_suyang, curve_siyang]):
        max_power = power.query.with_entities(func.max(power.total_power), power.date, power.number)\
            .filter(power.date >= st, power.date <= ed).first()
        max_weather = Weather.query.with_entities(Weather.temp, Weather.humidity, Weather.wind_speed, Weather.prec)\
            .filter(Weather.date == max_power[1], Weather.hour == max_power[2]).first()
        min_power = min(curve_data)
        mean_power = sum(curve_data)/len(curve_data)
        if max_weather is None:
            max_info.append([round(max_power[0], 1), get_time_str(max_power[1], max_power[2]), 0.0,
                             0.0, 0.0, 0.0, round(min_power, 1), round(mean_power, 1)])
        else:
            max_info.append([round(max_power[0], 1), get_time_str(max_power[1], max_power[2]), round(max_weather[0], 1),
                             round(max_weather[1], 1), round(max_weather[2], 1),
                             round(max_weather[3], 1), round(min_power, 1), round(mean_power, 1)])

    template_data = {'time': datetime.now().strftime('%Y/%m/%d\t%H:%M'),
                     'load': DATA['nowpower'],
                     'curve_time': curve_time,
                     'curve_sh': curve_sh,
                     'curve_pred_sh': curve_pred_sh,
                     'curve_sq': curve_sq,
                     'curve_suyang': curve_suyang,
                     'curve_siyang': curve_siyang,
                     'pie_time': curve_time[0],
                     'pie_data': pie_data,
                     'max_info': max_info,
                     }
    return template_data


@app.route('/hisload')
def hisload():
    # 查询 k-1 天各个区县的24h的负荷功率
    st = (datetime.now()-timedelta(days=1)).date()
    ed = (datetime.now()-timedelta(days=1)).date()
    return render_template('hisload.html', template_data=query_st_ed(st, ed))


@app.route('/query_data', methods=['POST'])
def query_data():
    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    return query_st_ed(start_date, end_date)


# init_db()
if __name__ == "__main__":
    db.create_all()
    # query_st_ed(st=(datetime.now() - timedelta(days=1)).date(), ed=(datetime.now() - timedelta(days=1)).date())
    if len(sys.argv) == 3:
        init_db_flag = sys.argv[1]
        delete_user_table_flag = sys.argv[2]
        print(f"参数设置: \n init_db_flag = {init_db_flag} \n delete_user_table_flag = {delete_user_table_flag}")
    else:
        init_db_flag = False
        delete_user_table_flag = False
        print("采用默认参数:\ninit_db_flag = False\ndelete_user_table_flag = False")
    if init_db_flag:
        init_db()
        logger.info("init_db() 成功")
        print("init_db() 成功 ")
    if delete_user_table_flag:
        delete_user_table()
        init_user()
        logger.info("init_user() 成功")
        print("init_user() 成功")

    # init_db()
    # delete_user_table()  # 删除并从 excel中重新读取所有的用户信息，部署时可注释
    if User.query.count() == 0:
        init_user()
    print("init_user() 成功")

    user_num_list = [User.query.count(), Usersuqian.query.count(), Usersuyang.query.count(), Usersiyang.query.count()]
    user_list = [User, Usersuqian, Usersuyang, Usersiyang]
    power_list = [Power, Powersuqian, Powersuyang, Powersiyang]
    totalpower_list = [Totalpower, Totalpowersuqian, Totalpowersuyang, Totalpowersiyang]

    init_power(user_num_list, user_list, power_list)  # 初始化k-2天的数据 全部为0  加权后的数据
    print("init_power() 成功")
    # 请求用户表 user_list 中用户的负荷  保存到 power_list   区域求和到 total_power
    getPowerData_threadingpool(user_num_list, user_list, power_list, totalpower_list)  # 请求获得K-1时刻的数据 异常采用k-2数据
    print("getPowerData_threadingpool() 成功")

    # init_predict_power  K-1时刻的数据  保存的预测结果数据初始化
    init_predict_power()
    print("init_predict_power() 成功")

    # 删除大于等于K-1的所有数据 请求初始化K-1时刻  getWeatherDataP() 获得K所有数据
    Weather.query.filter(Weather.date >= (datetime.now() - timedelta(days=1)).date()).delete()
    db.session.commit()
    init_weather()  # 请求获得K-1 K数据
    print("init_weather() 成功")

    getWeatherDataP()  # K 预测气象数据 [8, 9, 10, ...]
    print("getWeatherDataP() 成功")

    getWeatherDataN()  # 当前时刻气象数据
    print("getWeatherDataN() 成功")   # 天气表格DATA['weather_table_data']   天气曲线DATA["weather_curve"]

    delete_hst_data()
    print("delete_hst_data() 成功")

    predict_power()  # DATA["power_curve"] DATA['power_table_data'] DATA["weather_curve"] DATA['weather_table_data'] DATA['nowpower']
    print("predict_power() 成功")

    # 开启更新线程
    update_weather_thread = threading.Thread(target=update_weather_thread)
    update_weather_thread.daemon = True
    update_weather_thread.start()

    update_power_thread = threading.Thread(target=update_power_thread, args=(user_num_list, user_list, power_list, totalpower_list))
    update_power_thread.daemon = True
    update_power_thread.start()

    print("开启后台更新线程...")
    print("运行程序...")
    print("打开http地址: http://127.0.0.1:5000/ ")
    app.run(debug=False)
